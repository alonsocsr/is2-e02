from datetime import datetime
import profile
from django.utils import timezone
import pytest
from django.urls import reverse
from django.contrib.auth.models import Permission,Group
from categories.models import Categorias
from content.models import Contenido
from profiles.models import Profile, Suscripcion
import openpyxl
from io import BytesIO

@pytest.mark.django_db
def test_update_profile_get(client,setup):
    user, _ = setup

  
    permiso_editar_perfil = Permission.objects.get(codename='editar_perfil')
    user.user_permissions.add(permiso_editar_perfil)
    client.force_login(user)
    
    url=reverse('profile')
    response = client.get(url)
    
    assert response.status_code == 200
    assert 'testuser' in response.content.decode()
    
@pytest.mark.django_db
def test_update_profile_post(client, setup):
    
    user, _ = setup
    client.force_login(user)
    permiso_editar_perfil = Permission.objects.get(codename='editar_perfil')
    user.user_permissions.add(permiso_editar_perfil)

    url = reverse('profile')  
    data = {
        'first_name': 'Juan',
        'last_name': 'Perez',
        'username': 'newuser',
    }

   
    response = client.post(url, data)
 
    user.refresh_from_db()

    f"Nombre de usuario luego de la actualizaciÃ³n: {user.username}"
    assert response.status_code == 302  
    assert user.username == 'newuser'
    assert user.first_name == 'Juan'
    assert user.last_name == 'Perez'

@pytest.mark.django_db
def test_ver_historial_compras(client,setup):
    user, _ = setup

    pytestmark = pytest.mark.filterwarnings("error")
    permiso_historial = Permission.objects.get(codename='ver_historial_compras')
    user.user_permissions.add(permiso_historial)
    client.force_login(user)
    
    url=reverse('ver_historial_compras')
    response = client.get(url)
    
    assert response.status_code == 200

@pytest.mark.django_db
def test_likes(client,setup):
    user, _ = setup
    
     
    categoria = Categorias.objects.create(nombre_categoria="TestCategoria")
    contenido = Contenido.objects.create(
        titulo='Nuevo Contenido',
        resumen='Este es un resumen del contenido',
        cuerpo='<p>Este es el cuerpo del contenido</p>',
        categoria=categoria,
        estado='Publicado',
        activo=True,
        fecha_publicacion=timezone.now().date(),
        vigencia=timezone.now().date() + timezone.timedelta(days=30),
        slug='nuevo-contenido',
        usuario_editor=user,
        cantidad_likes=0
    )
    print(f'Cantidad de likes antes de la accion: {contenido.cantidad_likes}') 

    
    url = reverse('like', kwargs={'id': contenido.pk})
    response = client.post(url, follow=True)
    

    contenido.refresh_from_db() 
    print(f'Cantidad de likes luego de la accion: {contenido.cantidad_likes}')  
    assert response.status_code == 200  
    assert contenido.cantidad_likes>=1
@pytest.mark.django_db
def test_dislikes(client,setup):
    user, _ = setup
    
     
    categoria = Categorias.objects.create(nombre_categoria="TestCategoria")
    contenido = Contenido.objects.create(
        titulo='Nuevo Contenido',
        resumen='Este es un resumen del contenido',
        cuerpo='<p>Este es el cuerpo del contenido</p>',
        categoria=categoria,
        estado='Publicado',
        activo=True,
        fecha_publicacion=timezone.now().date(),
        vigencia=timezone.now().date() + timezone.timedelta(days=30),
        slug='nuevo-contenido',
        usuario_editor=user,
        cantidad_dislikes=0
    )
    print(f'Cantidad de dislikes antes de la accion: {contenido.cantidad_dislikes}') 

    
    url = reverse('dislike', kwargs={'id': contenido.pk})
    response = client.post(url, follow=True)
    

    contenido.refresh_from_db() 
    print(f'Cantidad de dislikes luego de la accion: {contenido.cantidad_dislikes}')  
    assert response.status_code == 200  
    assert contenido.cantidad_dislikes>=1
    
@pytest.mark.django_db
def test_eliminar_cuenta_view(client, setup):
   
    user, _ = setup

    client.login(username=user.username, password='12345')  
 
    response = client.post(reverse('eliminar_cuenta'), data={'password': '12345'}) 


    assert not Profile.objects.filter(user=user).exists(), "La cuenta no ha sido eliminada correctamente."
    assert response.status_code == 302  
    
@pytest.mark.django_db
def test_exportar_compras_xlsx(client,setup):
    user, _ = setup

    categoria = Categorias.objects.create(nombre_categoria="Test Categoria")
    Suscripcion.objects.create(
        profile=user.profile,
        categoria=categoria,
        fecha_pago=datetime.now(),
        monto=100,
        medio_pago='TC'
    )

    financiero_group = Group.objects.get(name='Financiero')
    user.groups.add(financiero_group)

    client.login(username=user.username, password='12345')  

    url = reverse('exportar_compras_xlsx') 
    response = client.get(url) 


    assert response.status_code == 200
    assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert response['Content-Disposition'] == 'attachment; filename=historial_compras.xlsx'


    workbook = openpyxl.load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet.title == 'Historial de Compras'
    headers = ['Usuario', 'Categoría', 'Fecha del Pago', 'Hora del Pago', 'Monto', 'Medio de Pago']
    assert [cell.value for cell in sheet[1]] == headers

    assert sheet.cell(row=2, column=1).value == 'testuser'
    assert sheet.cell(row=2, column=2).value == 'Test Categoria'